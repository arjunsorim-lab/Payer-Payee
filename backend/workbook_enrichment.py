"""Integrated-workbook repository for the scoped Member 360 money workflow.

When ``SAVINGS_WORKBOOK_PATH`` is configured this module is the only claims and
evidence source used by the affected API endpoints. It never falls back to CSV,
MongoDB, or bundled browser data.
"""

from __future__ import annotations

import json
import os
from dataclasses import dataclass
from datetime import date, datetime, timezone
from hashlib import sha256
from pathlib import Path
from threading import RLock

from openpyxl import load_workbook
from dotenv import load_dotenv

try:
    from .claim_mapper import build_member_documents, normalize_claim
except ImportError:
    from claim_mapper import build_member_documents, normalize_claim

load_dotenv()

CLAIMS_SHEET = os.getenv("CLAIMS_WORKSHEET_NAME", "837_Claims")
ELIGIBILITY_SHEET = os.getenv("ELIGIBILITY_WORKSHEET_NAME", "834_Eligibility")
REASON_LEGEND_SHEET = os.getenv("REASON_LEGEND_WORKSHEET_NAME", "Reason_Code_Legend")
FIELD_DICTIONARY_SHEET = os.getenv("FIELD_DICTIONARY_WORKSHEET_NAME", "New_Fields_Dictionary")
DATA_NOTES_SHEET = os.getenv("DATA_NOTES_WORKSHEET_NAME", "Data_Notes_READ_ME")
REQUIRED_SHEETS = {
    CLAIMS_SHEET,
    ELIGIBILITY_SHEET,
    REASON_LEGEND_SHEET,
    FIELD_DICTIONARY_SHEET,
    DATA_NOTES_SHEET,
}

CALCULATION_VERSION = "workbook-money-v1"
PREDICTION_VERSION = "workbook-peer-forecast-v1"
SAVINGS_VERSION = "workbook-opportunity-v1"
RAG_INDEX_VERSION = "workbook-faiss-hash-v1"
GROQ_PROMPT_VERSION = "workbook-layman-explanation-v2"

_CACHE: dict[tuple[str, int, int], "WorkbookDatabase"] = {}
_ACTIVE_HASH = ""
_LOCK = RLock()


def _clean_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _sheet_records(sheet):
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows)]
    except StopIteration:
        return [], []
    records = []
    for source_row, row in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in row):
            continue
        record = {
            header: _clean_value(value)
            for header, value in zip(headers, row)
            if header
        }
        record["_source_row"] = source_row
        records.append(record)
    return headers, records


def _canonical_json(value):
    return json.dumps(value, sort_keys=True, separators=(",", ":"), default=str)


def _is_yes(value):
    return str(value or "").strip().upper() in {"Y", "YES", "TRUE", "1"}


def normalize_claim_id(value):
    return "".join(character for character in str(value or "").upper() if character.isalnum())


@dataclass(frozen=True)
class WorkbookDatabase:
    path: Path
    claims: tuple[dict, ...]
    selectable_claims: tuple[dict, ...]
    historical_claims: tuple[dict, ...]
    eligibility_rows: tuple[dict, ...]
    reason_legend_rows: tuple[dict, ...]
    field_dictionary_rows: tuple[dict, ...]
    data_notes_rows: tuple[dict, ...]
    report: dict

    def __post_init__(self):
        object.__setattr__(
            self,
            "claims_by_id",
            {
                normalize_claim_id(claim.get("claimId")): claim
                for claim in self.claims
            },
        )
        object.__setattr__(
            self,
            "selectable_by_id",
            {
                normalize_claim_id(claim.get("claimId")): claim
                for claim in self.selectable_claims
            },
        )
        grouped = {}
        for claim in self.selectable_claims:
            grouped.setdefault(claim.get("memberId"), []).append(claim)
        object.__setattr__(
            self,
            "claims_by_member",
            {
                member_id: tuple(sorted(rows, key=lambda item: (item.get("dos", ""), item.get("claimId", "")), reverse=True))
                for member_id, rows in grouped.items()
            },
        )
        object.__setattr__(
            self,
            "members",
            tuple(sorted(build_member_documents(list(self.selectable_claims)), key=lambda item: (item.get("latestServiceDate", ""), item.get("memberId", "")), reverse=True)),
        )

    @property
    def workbook_hash(self):
        return self.report["workbook_hash"]

    def find_claim(self, claim_id, selectable_only=True):
        normalized = normalize_claim_id(claim_id)
        source = self.selectable_by_id if selectable_only else self.claims_by_id
        claim = source.get(normalized)
        if claim:
            return claim
        prefix = "".join(character for character in normalized if character.isalpha()).upper()
        digits = "".join(character for character in normalized if character.isdigit())
        if not prefix or not digits:
            return None
        numeric_id = int(digits)
        return next(
            (
                item for item in source.values()
                if (
                    "".join(
                        character
                        for character in normalize_claim_id(item.get("claimId"))
                        if character.isalpha()
                    ).upper()
                    == prefix
                    and int(
                        "".join(
                            character
                            for character in normalize_claim_id(item.get("claimId"))
                            if character.isdigit()
                        )
                    )
                    == numeric_id
                )
            ),
            None,
        )

    def member_claims(self, member_id):
        return list(self.claims_by_member.get(str(member_id or "").strip(), ()))

    def source_banner(self):
        return {
            "message": f"Workbook demonstration database active: {self.path.name}",
            "workbook_name": self.path.name,
            "workbook_hash": self.report["workbook_hash"],
            "workbook_hash_short": self.report["workbook_hash"][:12],
            "import_time": self.report["import_time"],
            "modified_time": self.report["modified_time"],
            "file_size": self.report["file_size"],
            "source_row_hash": self.report["source_row_hash"],
            "selectable_claim_count": self.report["selectable_claim_count"],
            "historical_reference_count": self.report["historical_reference_count"],
            "calculation_version": CALCULATION_VERSION,
            "prediction_version": PREDICTION_VERSION,
            "savings_version": SAVINGS_VERSION,
            "rag_index_version": RAG_INDEX_VERSION,
            "groq_prompt_version": GROQ_PROMPT_VERSION,
        }


def _notify_hash_change(previous_hash, next_hash):
    if not previous_hash or previous_hash == next_hash:
        return
    for module_name, function_name in (
        ("backend.financial_engine", "clear_financial_cache"),
        ("backend.workbook_rag", "clear_rag_cache"),
        ("backend.workbook_llm", "clear_llm_caches"),
    ):
        try:
            module = __import__(module_name, fromlist=[function_name])
            getattr(module, function_name)()
        except (ImportError, AttributeError):
            continue


def _load(path):
    workbook_path = Path(path).expanduser().resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Configured workbook cannot be loaded: {workbook_path}")

    stat = workbook_path.stat()
    workbook_hash = sha256(workbook_path.read_bytes()).hexdigest()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    missing = REQUIRED_SHEETS.difference(workbook.sheetnames)
    if missing:
        workbook.close()
        raise ValueError(
            "Configured workbook is missing required sheet(s): "
            + ", ".join(sorted(missing))
        )

    headers, claim_rows = _sheet_records(workbook[CLAIMS_SHEET])
    _, eligibility_rows = _sheet_records(workbook[ELIGIBILITY_SHEET])
    _, reason_rows = _sheet_records(workbook[REASON_LEGEND_SHEET])
    _, dictionary_rows = _sheet_records(workbook[FIELD_DICTIONARY_SHEET])
    _, notes_rows = _sheet_records(workbook[DATA_NOTES_SHEET])
    workbook.close()

    claims = []
    for row in claim_rows:
        raw_fields = {key: value for key, value in row.items() if key != "_source_row"}
        claim = normalize_claim(raw_fields)
        if not claim.get("claimId") or not claim.get("memberId"):
            continue
        historical = _is_yes(raw_fields.get("Is_Historical_Reference_Record"))
        claim.update({
            "workbookFields": raw_fields,
            "workbookSourceRow": row["_source_row"],
            "isHistoricalReference": historical,
            "episodeId": str(raw_fields.get("Episode_ID") or "").strip(),
            "sourceWorkbook": workbook_path.name,
            "sourceWorkbookHash": workbook_hash,
            "sourceRowHash": sha256(_canonical_json(raw_fields).encode("utf-8")).hexdigest(),
            "calculationVersion": CALCULATION_VERSION,
            "predictionVersion": PREDICTION_VERSION,
            "ragIndexVersion": RAG_INDEX_VERSION,
        })
        claims.append(claim)

    historical_claims = tuple(claim for claim in claims if claim["isHistoricalReference"])
    selectable_claims = tuple(claim for claim in claims if not claim["isHistoricalReference"])
    source_row_hash = sha256(
        _canonical_json([
            {"claim_id": claim["claimId"], "row_hash": claim["sourceRowHash"]}
            for claim in claims
        ]).encode("utf-8")
    ).hexdigest()
    report = {
        "source_workbook": workbook_path.name,
        "source_workbook_path": str(workbook_path),
        "workbook_hash": workbook_hash,
        "modified_time": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
        "file_size": stat.st_size,
        "source_row_hash": source_row_hash,
        "import_time": datetime.now(timezone.utc).isoformat(),
        "claims_sheet": CLAIMS_SHEET,
        "claim_columns": headers,
        "claim_column_count": len(headers),
        "total_claim_count": len(claims),
        "selectable_claim_count": len(selectable_claims),
        "historical_reference_count": len(historical_claims),
        "calculation_version": CALCULATION_VERSION,
        "prediction_version": PREDICTION_VERSION,
        "savings_version": SAVINGS_VERSION,
        "rag_index_version": RAG_INDEX_VERSION,
        "groq_prompt_version": GROQ_PROMPT_VERSION,
    }
    return WorkbookDatabase(
        path=workbook_path,
        claims=tuple(claims),
        selectable_claims=selectable_claims,
        historical_claims=historical_claims,
        eligibility_rows=tuple(eligibility_rows),
        reason_legend_rows=tuple(reason_rows),
        field_dictionary_rows=tuple(dictionary_rows),
        data_notes_rows=tuple(notes_rows),
        report=report,
    )


def load_workbook_database(path=None, *, force=False):
    global _ACTIVE_HASH
    configured = path or os.getenv("SAVINGS_WORKBOOK_PATH", "").strip()
    if not configured:
        raise RuntimeError("SAVINGS_WORKBOOK_PATH is required for the workbook demonstration database.")
    workbook_path = Path(configured).expanduser().resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Configured workbook cannot be loaded: {workbook_path}")
    stat = workbook_path.stat()
    key = (str(workbook_path), stat.st_mtime_ns, stat.st_size)
    with _LOCK:
        if not force and key in _CACHE:
            return _CACHE[key]
        database = _load(workbook_path)
        previous_hash = _ACTIVE_HASH
        _ACTIVE_HASH = database.workbook_hash
        _CACHE.clear()
        _CACHE[key] = database
        _notify_hash_change(previous_hash, database.workbook_hash)
        return database


def read_savings_workbook(path):
    """Compatibility wrapper returning every row plus integrated metadata."""
    database = load_workbook_database(path)
    return list(database.claims), dict(database.report)


def clear_workbook_cache():
    global _ACTIVE_HASH
    with _LOCK:
        _CACHE.clear()
        _ACTIVE_HASH = ""
