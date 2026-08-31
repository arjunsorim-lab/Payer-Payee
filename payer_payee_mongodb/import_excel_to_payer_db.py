"""Import EDI_834_837 Excel workbook into local MongoDB 'payer' database."""

import os
import sys
from pathlib import Path
from hashlib import sha256
from datetime import datetime, date
import openpyxl
from pymongo import MongoClient, ASCENDING, ReplaceOne

# Import normalizer from backend
project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from backend.claim_mapper import normalize_claim, clean_raw_row, build_member_documents

MONGODB_URI = os.getenv("MONGODB_URI", "mongodb://localhost:27017/")
MONGODB_DB = os.getenv("MONGODB_DB", "payer")

DEFAULT_EXCEL_PATH = r"C:\Users\lenovo\Downloads\EDI_834_837_20_members_ENRICHED (2).xlsx"


def clean_cell(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def read_sheet_records(sheet):
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows)]
    except StopIteration:
        return [], []
    records = []
    for source_row, row in enumerate(rows, start=2):
        if not any(v not in (None, "") for v in row):
            continue
        record = {
            header: clean_cell(value)
            for header, value in zip(headers, row)
            if header
        }
        record["_source_row"] = source_row
        records.append(record)
    return headers, records


def import_excel(excel_path_str: str, mongo_uri: str = MONGODB_URI, db_name: str = MONGODB_DB):
    excel_path = Path(excel_path_str).resolve()
    if not excel_path.is_file():
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")

    print(f"[*] Reading Excel workbook: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    client = MongoClient(mongo_uri, serverSelectionTimeoutMS=5000)
    client.admin.command("ping")
    db = client[db_name]
    print(f"[*] Connected to MongoDB '{mongo_uri}', target database: '{db_name}'")

    # 1. 837_Claims
    claims_sheet = wb["837_Claims"] if "837_Claims" in wb.sheetnames else None
    if not claims_sheet:
        raise ValueError("Sheet '837_Claims' not found in workbook")

    claim_headers, raw_claims = read_sheet_records(claims_sheet)
    print(f"[*] Extracted {len(raw_claims)} raw claim rows from '837_Claims'")

    # Normalize claims for application consumption
    normalized_claims = []
    for row in raw_claims:
        raw_cleaned = clean_raw_row(row)
        norm = normalize_claim(raw_cleaned)
        if norm.get("claimId") and norm.get("memberId"):
            # Include raw row fields for detailed viewing
            norm["raw"] = raw_cleaned
            normalized_claims.append(norm)

    print(f"[*] Normalized {len(normalized_claims)} valid claims")

    # Build member documents from normalized claims
    member_docs = build_member_documents(normalized_claims)
    print(f"[*] Generated {len(member_docs)} member summaries")

    # 2. 834_Eligibility
    eligibility_records = []
    if "834_Eligibility" in wb.sheetnames:
        _, eligibility_records = read_sheet_records(wb["834_Eligibility"])
        print(f"[*] Extracted {len(eligibility_records)} eligibility records")

    # 3. Reason_Code_Legend
    reason_records = []
    if "Reason_Code_Legend" in wb.sheetnames:
        _, reason_records = read_sheet_records(wb["Reason_Code_Legend"])
        print(f"[*] Extracted {len(reason_records)} reason code records")

    # 4. New_Fields_Dictionary
    field_dict_records = []
    if "New_Fields_Dictionary" in wb.sheetnames:
        _, field_dict_records = read_sheet_records(wb["New_Fields_Dictionary"])
        print(f"[*] Extracted {len(field_dict_records)} field dictionary records")

    wb.close()

    # Bulk Upsert to MongoDB
    print(f"[*] Writing collections into database '{db_name}'...")

    # claims collection
    if normalized_claims:
        db.claims.create_index("claimId", unique=True)
        db.claims.create_index([("memberId", 1), ("diagnosisCode", 1), ("dos", -1)])
        db.claims.create_index("dos")
        db.claims.bulk_write(
            [ReplaceOne({"claimId": doc["claimId"]}, doc, upsert=True) for doc in normalized_claims],
            ordered=False
        )

    # members collection
    if member_docs:
        db.members.create_index("memberId", unique=True)
        db.members.bulk_write(
            [ReplaceOne({"memberId": doc["memberId"]}, doc, upsert=True) for doc in member_docs],
            ordered=False
        )

    # raw 837_claims collection
    if raw_claims:
        db["837_claims"].create_index("Claim_ID", unique=True)
        db["837_claims"].create_index("Member_ID")
        db["837_claims"].bulk_write(
            [ReplaceOne({"Claim_ID": doc["Claim_ID"]}, doc, upsert=True) for doc in raw_claims],
            ordered=False
        )

    # eligibility collection
    if eligibility_records:
        db["834_eligibility"].drop()
        db["834_eligibility"].insert_many(eligibility_records, ordered=False)

    # reason codes
    if reason_records:
        db["reason_codes"].drop()
        db["reason_codes"].insert_many(reason_records, ordered=False)

    # metadata
    db["import_metadata"].replace_one(
        {"_id": "latest_import"},
        {
            "_id": "latest_import",
            "source_file": str(excel_path),
            "file_size": excel_path.stat().st_size,
            "imported_at": datetime.now().isoformat(),
            "counts": {
                "claims": len(normalized_claims),
                "members": len(member_docs),
                "eligibility": len(eligibility_records),
                "reason_codes": len(reason_records),
            }
        },
        upsert=True
    )

    print("\n[✓] Successfully imported Excel data into MongoDB database '{0}':".format(db_name))
    print(f"    - claims collection:        {db.claims.count_documents({})} documents")
    print(f"    - members collection:       {db.members.count_documents({})} documents")
    print(f"    - 837_claims collection:    {db['837_claims'].count_documents({})} documents")
    print(f"    - 834_eligibility:          {db['834_eligibility'].count_documents({})} documents")
    print(f"    - reason_codes:             {db['reason_codes'].count_documents({})} documents")


if __name__ == "__main__":
    path_arg = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL_PATH
    import_excel(path_arg)
